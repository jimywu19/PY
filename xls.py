from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

wb = load_workbook(filename='sample.xlsx')

ws = wb.active
'''
ws['A1'] = 43
a = []
for i in range(15):
   a.append(i)
   ws.append(a)
wb.save('sample.xlsx')
'''

for row in ws.iter_rows(min_row=1,max_row=5,min_col=2,max_col=4):
    print(row)